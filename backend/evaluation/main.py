import json
import os
import re
import time
import uuid
from neo4j import GraphDatabase
from openpyxl import Workbook
import requests
import pandas as pd
from config import ANSWER_QUESTION, FILE_SERVICE_BASE_URL, FILES_DIR, NEO4J_PASSWORD, NEO4J_URI, NEO4J_USERNAME, REGISTER_URL, SESSION_URL
from llm import LLM 

llm_instance = LLM.get_instance()
llm = llm_instance.llm


payload_register = {
    "email": "email@email.com",
    "name": "name",
    "password": "password",
    "profesion": "estudiante"
}
response = requests.post(REGISTER_URL, json=payload_register)


payload_session = {
    "user_id": "email@email.com",
    "session_id": str(uuid.uuid4()), 
    "session_name": "Test Chat"
}
response = requests.post(SESSION_URL, json=payload_session)
session_id = response.json()["session_id"]

def upload_file(path, session_id):
    url = f"{FILE_SERVICE_BASE_URL}/files/upload"
    with open(path, "rb") as f:
        files = {"file": f}
        data = {"session_id": session_id}
        r = requests.post(url, files=files, data=data)
    if r.status_code != 200:
        raise requests.HTTPError(f"Upload failed for {path}: {r.text}", response=r)
    result = r.json()
    print(f"Uploaded {os.path.basename(path)} -> file_id={result.get('file_id')}")
    return result.get("file_id")


file_paths = [
    os.path.join(FILES_DIR, f)
    for f in os.listdir(FILES_DIR)
    if f.lower().endswith(".pdf")
]

file_ids = [upload_file(p, session_id) for p in file_paths]

def get_file_status(file_id):
    url = f"{FILE_SERVICE_BASE_URL}/files/{file_id}"
    r = requests.get(url)
    if r.status_code != 200:
        raise Exception(f"Failed to get status for {file_id}: {r.text}")
    return r.json().get("status")

def wait_for_processing(file_ids, interval=20):
    """Polls file statuses every `interval` seconds."""
    print("\nWaiting for files to be processed...")
    processed = set()
    while len(processed) < len(file_ids):
        for fid in file_ids:
            if fid in processed:
                continue
            try:
                status = get_file_status(fid)
                if status == "processed":
                    print(f"File {fid}: {status}")
                    processed.add(fid)
                elif status == "error":
                    print(f"File {fid} failed to process. Stopping.")
                    return False
            except Exception as e:
                print(f"Error checking {fid}: {e}")
        if len(processed) < len(file_ids):
            time.sleep(interval)
    print("All files processed successfully.")
    return True

wait_for_processing(file_ids)

# ==================== NEO4J HELPER ====================
class Neo4jHelper:
    def __init__(self, uri, user, password):
        self.driver = GraphDatabase.driver(uri, auth=(user, password))

    def close(self):
        self.driver.close()

    def get_node_by_id(self, node_id):
        query = f"MATCH (n) WHERE elementId(n) = \"{node_id}\" RETURN n" 
        with self.driver.session() as session:
            result = session.run(query)
            record = result.single()
            if record:
                node = record["n"]  
                if node and node.get("text"): 
                    return node["text"]
            return None

neo4j_helper = Neo4jHelper(NEO4J_URI, NEO4J_USERNAME, NEO4J_PASSWORD)

# ==================== QUERY ====================
def query_chatbot(question, session_id):
    payload = {
        "query": question,
        "session_id": session_id
    }
    try:
        r = requests.post(ANSWER_QUESTION, json=payload)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f"Error querying: {e}")
        return None

# ===============================================


file_path = os.path.join(FILES_DIR, "questions_answers.xlsx")
df = pd.read_excel(file_path)
dataset = df.to_dict(orient="records")

results = []

for item in dataset:
    question = item.get("Question")
    ground_truth = item.get("Answer")

    resp = query_chatbot(question, session_id)
    if not resp:
        continue

    chatbot_answer = resp.get("answer")
    retriever_results = resp.get("retriever_result", [])
    answer_created_at = resp.get("answer_created_at")

    node_contexts = []
    for rr in retriever_results:
        list_ids = rr.get("listIds", [])
        if list_ids:
            node_id = list_ids[0]
            node_data = neo4j_helper.get_node_by_id(node_id)
            if node_data:
                node_contexts.append(node_data)

    result_entry = {
        "question": question,
        "ground_truth": ground_truth,
        "new_answer": chatbot_answer,
        "context": node_contexts
    }
    results.append(result_entry)

wb = Workbook()
ws = wb.active
ws.append(["user_input","reference","response","retrieved_contexts","relevance","correctness","fluency","context_relevance"])
for result in results:
    query = result["question"]
    generated = result["new_answer"]
    reference = result["ground_truth"]
    context_raw = result["context"]
    if isinstance(context_raw, list):
        if not context_raw:  
            context_value = "None"
        else:  
            context_value = "\n".join(item.replace("\n", " ") for item in context_raw)
    else:
        context_value = str(context_raw).replace("\n", " ")
    prompt = f"""
    You are an evaluator. Rate the generated answer based on the query.

    Query: {query}
    Generated answer: {generated}
    Reference answer: {reference if reference else 'N/A'}
    Retrived context:{context_value}

    Provide a JSON with scores between 0 and 1 for:
    - relevance: how relevant the answer is to the query
    - correctness: how factually correct the answer is
    - fluency: how well-written the answer is
    - context relevance: how relevant is the context retrieved
    """
    llm_response = llm.invoke(prompt)
    llm_response_content = llm_response.content
    match = re.search(
        r"```(?:json)?\s*(\{.*?\})\s*```",
        llm_response_content,
        re.DOTALL | re.IGNORECASE
    )
    data = None
    if match:
        json_str = match.group(1)
        try:
            data = json.loads(json_str)
        except json.JSONDecodeError as e:
            print(f"Invalid JSON format: {e}")
    else:
        print("No JSON found in response.")
    relevance          = data.get("relevance", 0.0)          if data else 0.0
    correctness        = data.get("correctness", 0.0)        if data else 0.0
    fluency            = data.get("fluency", 0.0)            if data else 0.0
    context_relevance  = data.get("context_relevance", 0.0)  if data else 0.0
    ws.append([query,reference,generated,context_value,relevance,correctness,fluency,context_relevance])

output_path = os.path.join(FILES_DIR, "metrics.xlsx")
wb.save(output_path)

total_queries = len(results)
if total_queries > 0:
    avg_relevance = sum(ws.cell(row=i, column=5).value for i in range(2, total_queries+2)) / total_queries
    avg_correctness = sum(ws.cell(row=i, column=6).value for i in range(2, total_queries+2)) / total_queries
    avg_fluency = sum(ws.cell(row=i, column=7).value for i in range(2, total_queries+2)) / total_queries
    avg_context_relevance = sum(ws.cell(row=i, column=8).value for i in range(2, total_queries+2)) / total_queries
else:
    avg_relevance = avg_correctness = avg_fluency = avg_context_relevance = 0.0

print(f"Queries done: {total_queries}")
print(f"Avg relevance: {avg_relevance:.3f}")
print(f"Avg correctness: {avg_correctness:.3f}")
print(f"Avg fluency: {avg_fluency:.3f}")
print(f"Avg context relevance: {avg_context_relevance:.3f}")
